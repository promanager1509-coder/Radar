#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PSNHUB — Конвертер Excel → JSON
================================
Поддерживает разные структуры колонок у разных застройщиков.
Добавить нового застройщика = добавить его маппинг в DEVELOPER_MAPS.

Использование:
    python excel_to_json.py

Файлы Excel кладёшь в:  tools\source_excel\
Готовые JSON появятся в: data\developers\
"""

import openpyxl
import json
import os
import re
import sys
from datetime import datetime

# ─────────────────────────────────────────────────────────────
# ПУТИ (относительно папки где лежит скрипт)
# ─────────────────────────────────────────────────────────────
SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
SOURCE_DIR   = os.path.join(SCRIPT_DIR, "source_excel")
OUTPUT_DIR   = os.path.join(SCRIPT_DIR, "..", "data", "developers")

# ─────────────────────────────────────────────────────────────
# МАППИНГИ КОЛОНОК ПО ЗАСТРОЙЩИКАМ
# Ключ = подстрока в имени файла (нижний регистр)
# Значение = словарь {наше_поле: название_колонки_в_excel}
# ─────────────────────────────────────────────────────────────
DEVELOPER_MAPS = {

    # ── ГК ПИК ──────────────────────────────────────────────
    "пик": {
        "slug":        "pik",
        "developer":   "ГК ПИК",
        "deal":        "sale",  # по умолчанию продажа
        "col": {
            "id":           "Номер_помещения",
            "jk":           "ЖК",
            "building":     "Дом",
            "type":         "Тип_объекта",
            "district":     "АО Москвы",
            "city":         "Регион",
            "address":      "Адрес",
            "area":         "Площадь_м2",
            "delivery":     "Срок_сдачи",
            "finishing":    "Отделка",
            "price":        "Цена_базовая_руб",
            "price_sale":   "Цена_спецпредложение_руб",
            "url_developer":"ссылка на объект",
            "rent_month":   "ориентировочный доход от аренды в мес ",
            "commission":   "комиссия агента",
        }
    },

    # ── А101 аренда ──────────────────────────────────────────
    "а101": {
        "slug":        "a101",
        "developer":   "А101",
        "deal":        "auto",  # берём из колонки Тип_сделки
        "col": {
            "id":           "ID (уникальный_код_объекта)",
            "jk":           "ЖК (название_проекта)",
            "building":     "Корпус (номер_дома)",
            "type":         "Тип_объекта (psn/office)",
            "deal_col":     "Тип_сделки (rent/sale)",
            "format":       "Формат_объекта (standard/gab_ready/gab_franchise)",
            "floor":        "Этаж (номер)",
            "district":     "Округ_Район (админ_локация)",
            "city":         "Город (moscow/mo/...)",
            "address":      "Адрес (полный_почтовый)",
            "area":         "Площадь_м2 (число)",
            "price_rent":   "Стоимость_в_месяц_руб (для_rent)",
            "price":        "Цена_продажи_руб (для_sale)",
            "status":       "Статус_дома (сдан/строится)",
            "metro":        "Метро_1 (ближайшее)",
            "metro2":       "Метро_2 (второе_метро)",
            "url_developer":"Официальная_ссылка (URL_застройщика)",
            "url_3d":       "3D  тур по ЖК ",
        }
    },

    # ── ЛСР ──────────────────────────────────────────────────
    "лср": {
        "slug":        "lsr",
        "developer":   "ГК ЛСР",
        "deal":        "sale",
        "col": {
            "id":           "Номер_помещения",
            "jk":           "Жилой_комплекс",
            "building":     "Корпус",
            "type":         "Тип_объекта",
            "floor":        "Этаж",
            "district":     "Район",
            "city":         "Город",
            "address":      "Адрес",
            "metro":        "Метро",
            "metro_min":    "Минут_до_метро",
            "area":         "Площадь_м2",
            "ceiling":      "Высота_потолков",
            "power":        "Мощность_кВт",
            "finishing":    "Отделка",
            "delivery":     "Срок_сдачи_Готовность",
            "price":        "Цена_руб",
            "price_sale":   "спецпредложение",
            "commission":   "Комиссия_%",
            "url_developer":"Источник_URL",
            "url_3d":       "пешеходный тур 360 градусов ",
        }
    },

    # ── ШАБЛОН ДЛЯ НОВОГО ЗАСТРОЙЩИКА ────────────────────────
    # Скопируй этот блок, замени ключ и заполни col{}
    # "самолёт": {
    #     "slug":      "samolet",
    #     "developer": "Самолёт",
    #     "deal":      "sale",
    #     "col": {
    #         "id":    "Номер",
    #         "jk":    "ЖК",
    #         ...
    #     }
    # },
}

# ─────────────────────────────────────────────────────────────
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ─────────────────────────────────────────────────────────────

def clean_price(val):
    """41 203 240 руб. → 41203240"""
    if val is None:
        return 0
    s = str(val).replace(" ", "").replace("\xa0", "")
    s = re.sub(r'[^\d.,]', '', s)
    s = s.replace(",", ".")
    parts = s.split(".")
    if len(parts) > 1:
        s = parts[0]
    try:
        return int(float(s))
    except:
        return 0

def clean_area(val):
    """100.3 м² → 100.3"""
    if val is None:
        return 0.0
    s = str(val).replace(" ", "").replace("\xa0", "").replace("м²", "").replace("м2", "")
    s = re.sub(r'[^\d.,]', '', s)
    s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except:
        return 0.0

def clean_str(val):
    """Чистит строку от пробелов и None"""
    if val is None:
        return ""
    return str(val).strip()

def clean_floor(val):
    """'1.0' → 1"""
    if val is None:
        return None
    try:
        return int(float(str(val)))
    except:
        return None

def normalize_type(val):
    """Нормализует тип помещения"""
    v = clean_str(val).lower()
    if "офис" in v or "office" in v:
        return "Офис"
    if "габ" in v or "gab" in v or "ready" in v:
        return "ГАБ"
    if "пвз" in v or "pvz" in v or "пункт" in v:
        return "ПВЗ"
    if "франш" in v or "franchise" in v:
        return "ГАБ"
    if "премиум" in v or "premium" in v or "элит" in v:
        return "Премиум"
    return "ПСН"

def normalize_deal(val):
    """'аренда ' → 'rent', 'продажа' → 'sale'"""
    v = clean_str(val).lower()
    if "аренд" in v or "rent" in v:
        return "rent"
    return "sale"

def normalize_city(val):
    """'Москва ' → 'Москва'"""
    v = clean_str(val)
    if "москва" in v.lower() or "moscow" in v.lower():
        return "Москва"
    return v

def normalize_delivery(val):
    """'до 28 апреля 2028' / '2026.0' → '2028-Q2' / '2026'"""
    if val is None:
        return ""
    s = str(val).strip()
    # Если просто год: 2026.0
    year_match = re.search(r'(202\d)', s)
    if year_match:
        year = year_match.group(1)
        # Попробуем найти квартал по месяцу
        months_q = {
            "январ": "Q1", "феврал": "Q1", "март": "Q1",
            "апрел": "Q2", "май": "Q2", "мая": "Q2", "июн": "Q2",
            "июл": "Q3", "август": "Q3", "сентябр": "Q3",
            "октябр": "Q4", "ноябр": "Q4", "декабр": "Q4",
        }
        for month, q in months_q.items():
            if month in s.lower():
                return f"{year}-{q}"
        return year
    return s

def split_metro(val):
    """'ЗИЛ, МЦК ЗИЛ, Тульская, Автозаводская' → ['Тульская', 'Автозаводская']"""
    if not val:
        return []
    parts = [p.strip() for p in str(val).split(",")]
    # Фильтруем МЦК и слишком короткие
    result = []
    for p in parts:
        if p and len(p) > 3 and "МЦК" not in p:
            result.append(p)
    return result[:2]  # максимум 2 станции

def normalize_commission(val):
    """'3%' / '3.5' / None → 3.5"""
    if val is None:
        return 0
    s = str(val).replace("%", "").replace(",", ".").strip()
    try:
        return round(float(s), 1)
    except:
        return 0

def make_id(slug, val, row_num):
    """Генерирует id если в файле нет уникального"""
    clean = re.sub(r'[^a-zA-Z0-9А-Яа-яЁё]', '-', str(val or row_num))
    clean = clean.strip('-')[:30]
    return f"{slug}-{clean}" if clean else f"{slug}-{row_num}"

# ─────────────────────────────────────────────────────────────
# ОСНОВНАЯ ФУНКЦИЯ КОНВЕРТАЦИИ
# ─────────────────────────────────────────────────────────────

def get_col_index(headers, name):
    """Находит индекс колонки по имени (нечувствительно к пробелам)"""
    name_clean = name.strip().lower()
    for i, h in enumerate(headers):
        if h and h.strip().lower() == name_clean:
            return i
    return None

def convert_file(filepath, dev_map):
    """Конвертирует один Excel файл в список объектов"""
    slug      = dev_map["slug"]
    developer = dev_map["developer"]
    deal_def  = dev_map["deal"]
    col_map   = dev_map["col"]

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Читаем заголовки (строка 1)
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        headers.append(str(v).strip() if v else "")

    # Строим индекс колонок
    idx = {}
    for field, col_name in col_map.items():
        i = get_col_index(headers, col_name)
        if i is not None:
            idx[field] = i
        else:
            print(f"  ⚠️  Колонка не найдена: '{col_name}' (поле {field}) — будет пустым")

    units = []
    skipped = 0
    seen_ids = set()

    for row_num in range(2, ws.max_row + 1):
        def cell(field):
            i = idx.get(field)
            return ws.cell(row_num, i + 1).value if i is not None else None

        # Пропускаем пустые строки
        area = clean_area(cell("area"))
        if area <= 0:
            skipped += 1
            continue

        # ID
        raw_id = clean_str(cell("id"))
        uid = make_id(slug, raw_id, row_num)
        # Дедупликация
        if uid in seen_ids:
            uid = f"{uid}-{row_num}"
        seen_ids.add(uid)

        # Тип сделки
        if deal_def == "auto":
            deal = normalize_deal(cell("deal_col"))
        else:
            deal = deal_def

        # Тип помещения
        unit_type = normalize_type(cell("type"))

        # Цена
        if deal == "rent":
            price = clean_price(cell("price_rent") or cell("price"))
        else:
            price = clean_price(cell("price_sale") or cell("price"))

        # Метро — объединяем metro + metro2
        metro_list = []
        m1 = clean_str(cell("metro"))
        if m1:
            metro_list += split_metro(m1)
        m2 = clean_str(cell("metro2"))
        if m2 and m2 not in metro_list:
            metro_list += split_metro(m2)
        metro_list = list(dict.fromkeys(metro_list))[:3]  # уникальные, макс 3

        # 3D тур
        url_3d = clean_str(cell("url_3d"))
        has_3d = bool(url_3d and url_3d.startswith("http"))

        unit = {
            "id":            uid,
            "jk":            clean_str(cell("jk")),
            "developer":     developer,
            "type":          unit_type,
            "deal":          deal,
            "price":         price,
            "area":          area,
            "floor":         clean_floor(cell("floor")),
            "finishing":     clean_str(cell("finishing")),
            "delivery":      normalize_delivery(cell("delivery")),
            "district":      clean_str(cell("district")),
            "city":          normalize_city(cell("city")),
            "metro":         metro_list,
            "address":       clean_str(cell("address"))[:100],
            "url_developer": clean_str(cell("url_developer")),
            "has_3d":        has_3d,
            "url_3d":        url_3d if has_3d else "",
            "commission":    normalize_commission(cell("commission")),
            "comment":       "",
        }
        units.append(unit)

    print(f"  ✅ Конвертировано: {len(units)} объектов (пропущено пустых: {skipped})")
    return units

# ─────────────────────────────────────────────────────────────
# ЗАПИСЬ JSON
# ─────────────────────────────────────────────────────────────

def save_json(units, slug, deal, output_dir):
    """Сохраняет JSON в нужную папку"""
    folder = os.path.join(output_dir, slug)
    os.makedirs(folder, exist_ok=True)
    filename = f"{slug}_{deal}.json"
    filepath = os.path.join(folder, filename)

    # Берём developer из первого объекта
    developer = units[0]["developer"] if units else slug

    output = {
        "developer": developer,
        "slug": slug,
        "updated": datetime.now().strftime("%Y-%m-%d"),
        "deal": deal,
        "units": units
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  💾 Сохранён: {filepath}")
    return filepath

# ─────────────────────────────────────────────────────────────
# ПОИСК МАППИНГА ПО ИМЕНИ ФАЙЛА
# ─────────────────────────────────────────────────────────────

def find_map(filename):
    """Находит маппинг по подстроке в имени файла"""
    name_lower = filename.lower()
    for key, dev_map in DEVELOPER_MAPS.items():
        if key in name_lower:
            return dev_map
    return None

# ─────────────────────────────────────────────────────────────
# ГЛАВНАЯ ФУНКЦИЯ
# ─────────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("PSNHUB — Конвертер Excel → JSON")
    print("=" * 55)

    # Проверяем папку с Excel
    if not os.path.exists(SOURCE_DIR):
        os.makedirs(SOURCE_DIR)
        print(f"\n📁 Создана папка: {SOURCE_DIR}")
        print(f"   Положи Excel файлы туда и запусти снова.\n")
        return

    # Ищем все Excel файлы
    xlsx_files = [
        f for f in os.listdir(SOURCE_DIR)
        if f.lower().endswith(".xlsx") and not f.startswith("~")
    ]

    if not xlsx_files:
        print(f"\n⚠️  Excel файлы не найдены в: {SOURCE_DIR}")
        print(f"   Положи .xlsx файлы в эту папку и запусти снова.\n")
        return

    print(f"\nНайдено файлов: {len(xlsx_files)}\n")

    success = 0
    errors  = 0

    for filename in xlsx_files:
        filepath = os.path.join(SOURCE_DIR, filename)
        print(f"📄 Обрабатываю: {filename}")

        dev_map = find_map(filename)
        if not dev_map:
            print(f"  ❌ Не найден маппинг для этого файла.")
            print(f"     Добавь застройщика в DEVELOPER_MAPS в скрипте.")
            print(f"     Ключ должен быть подстрокой имени файла (нижний регистр).")
            errors += 1
            print()
            continue

        try:
            units = convert_file(filepath, dev_map)
            if not units:
                print(f"  ⚠️  Объектов не найдено — проверь файл.")
                errors += 1
                print()
                continue

            # Разделяем продажу и аренду если в одном файле
            sale_units = [u for u in units if u["deal"] == "sale"]
            rent_units = [u for u in units if u["deal"] == "rent"]

            if sale_units:
                save_json(sale_units, dev_map["slug"], "sale", OUTPUT_DIR)
            if rent_units:
                save_json(rent_units, dev_map["slug"], "rent", OUTPUT_DIR)
            if not sale_units and not rent_units:
                save_json(units, dev_map["slug"], dev_map["deal"], OUTPUT_DIR)

            success += 1

        except Exception as e:
            print(f"  ❌ Ошибка: {e}")
            import traceback
            traceback.print_exc()
            errors += 1

        print()

    print("=" * 55)
    print(f"Готово: ✅ {success} файлов  |  ❌ {errors} ошибок")
    print(f"JSON файлы в: {OUTPUT_DIR}")
    print("=" * 55)
    print("\nСледующий шаг:")
    print("  cd C:\\Users\\user\\Radar")
    print("  git add .")
    print('  git commit -m "Обновление данных"')
    print("  git push")
    print()

if __name__ == "__main__":
    main()
